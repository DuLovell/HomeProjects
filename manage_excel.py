#!/usr/bin/env python3

from openpyxl import Workbook
import openpyxl
from os import listdir
import os


def create_new_xlsx():
    workbook = Workbook()
    name = input("Введите название документа:\n")
    return workbook, name

def display_all_xlsx():
    count = 1
    # Находим все существующие excel документы в рабочей директории в нужном нам формате
    # и выводим текстом на экран
    onlyfiles = [f.replace('.xlsx', '') for f in listdir(os.getcwd()) if f.endswith(".xlsx")]
    print("Выберите документ:")
    for xlsx in onlyfiles:
        print(f"{count}. {xlsx}")
        count += 1
    print("\n")
    choice = int(input())
    # Возвращаем объект workbook (без выбора листа для работы)
    workbook = openpyxl.load_workbook(onlyfiles[choice - 1] + ".xlsx")
    name = onlyfiles[choice - 1]
    return workbook, name

def manage_input(sheet):
    choice = int(input("""Выберите способ ввода данных:
    1. Из текстового файла
    2. Вручную с клавиатуры\n"""))
    if choice == 1:
        count = 1
        print("Выберите файл:")
        onlytxt = [f.replace('.txt', '') for f in listdir(os.getcwd()) if f.endswith(".txt")]
        for txt in onlytxt:
            print(f"{count}. {txt}")
            count += 1
        print("\n")
        txt_choice = int(input())
        file = open(onlytxt[txt_choice - 1] + ".txt", encoding='utf-8')
        file_lines = file.readlines()
        for row in file_lines:
            sheet.append([i.strip() for i in row.split(',')])
    elif choice == 2:
        print("Вводите текст построчно")
        #TODO Показать названия каждого столбца (1 ряд в таблице), если лист дополняется
        #TODO Иначе предупредить о вводе в первый ряд названия столбцов
        print("Введите 0, если хотите прекратить ввод текста")
        text = [i.strip() for i in input().split(',')]
        while text != ['0']:
            sheet.append(text)
            text = [i.strip() for i in input().split(',')]
    print("Таблица успешно заполнена!")
    return

def main():
    option = int(input("""Выберите действие (цифру):
1. Создать новый документ
2. Дополнить существующий документ\n"""))
    if option == 1:
        ## Создаем новый .xlsx документ
        wb, name = create_new_xlsx()  # должна вернуть объект документа + его имя (без .xlsx)
        # Определяем рабочий лист
        sheet = wb.active  # Выбираем единственную страницу в новом документе
        sheet.title = input("Введите название листа:\n")
    elif option == 2:
        # Предлагаем выбор нужного документа
        wb, name = display_all_xlsx()  # должна вернуть объект документа + его имя (без .xlsx)
        # Предлагаем выбор нужного рабочего листа
        count = 1
        print("Выберите рабочий лист:\n")
        print("0. Создать новый рабочий лист")
        for sheetname in wb.sheetnames:
            print(f"{count}. {sheetname}")
            count += 1
        choice = int(input())
        if choice == 0:
            sheet = wb.create_sheet(input("Введите название листа:\n"))
        # Определяем рабочий лист
        else:
            sheet = wb[wb.sheetnames[choice - 1]]
    ## Предлагаем выбор ввода текста 1 - из файла, 2 - с клавиатуры
    ## Заполняем .xlsx документ введенными данными
    manage_input(sheet)
    ## Сохраняем .xlsx документ
    wb.save(filename=(name + '.xlsx'))

main()
